import os

import markdown
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def extract_headers_and_content_sequential(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        md_content = f.read()

    html = markdown.markdown(md_content)
    soup = BeautifulSoup(html, 'html.parser')

    headers_content = []
    header_tags = ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']

    for element in soup.find_all(header_tags):
        header_level = element.name
        header_text = element.text.strip()

        content = []
        next_sibling = element.find_next_sibling()
        while next_sibling and next_sibling.name not in header_tags:
            content.append(next_sibling.text.strip())
            next_sibling = next_sibling.find_next_sibling()

        full_text = f"**{header_text} ({header_level.upper()})**\n" + "\n".join(content)
        headers_content.append(full_text)

    return headers_content


def process_folder(folder_path, output_file):
    data = []

    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.md'):
                file_path = os.path.join(root, file)
                headers_content = extract_headers_and_content_sequential(file_path)

                row = {'Filename': file}
                for idx, header_content in enumerate(headers_content):
                    row[f'Header {idx + 1}'] = header_content
                    row[f'Comment {idx + 1}'] = ""
                data.append(row)

    df = pd.DataFrame(data)
    max_headers = max((len(row) - 1) // 2 for row in data)

    column_order = ['Filename'] + [
        f'{label} {i + 1}'
        for i in range(max_headers)
        for label in ['Header', 'Comment']
    ]
    df = df.reindex(columns=column_order, fill_value='')

    wb = Workbook()
    ws = wb.active

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws.append(row)
        if r_idx > 0:
            for c_idx, cell_value in enumerate(row[1:], start=2):
                if isinstance(cell_value, str) and 'Header' in ws.cell(1, c_idx).value:
                    header_text, content_text = (
                        cell_value.split('\n', 1)
                        if '\n' in cell_value
                        else (cell_value, '')
                    )
                    cell_obj = ws.cell(row=r_idx + 1, column=c_idx)
                    cell_obj.value = f"{header_text}\n{content_text}"
                    cell_obj.alignment = Alignment(wrap_text=True)

    wb.save(output_file)


if __name__ == '__main__':
    folder_path = '../readme_files/'
    output_file = 'markdown_headers_with_comments.xlsx'
    process_folder(folder_path, output_file)

    print(f'Headers and content extracted and saved to {output_file}.')
